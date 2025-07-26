# -*- coding: utf-8 -*-

import base64
import io
import logging
from datetime import datetime, date
from odoo import models, fields, api
from odoo.exceptions import UserError

_logger = logging.getLogger('budget.wizard')

try:
    import openpyxl
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


class InvestmentBudgetWizard(models.TransientModel):
    """Мастер работы с инвестиционными бюджетами в формате ГК Хлебодар"""
    _name = 'investment.budget.wizard'
    _description = 'Импорт/Экспорт инвестиционных бюджетов'

    operation_type = fields.Selection([
        ('import', 'Импорт из Excel'),
        ('export', 'Экспорт в Excel'),
        ('sync', 'Синхронизация')
    ], 'Тип операции', required=True, default='import')

    # Параметры импорта
    import_file = fields.Binary('Excel файл для импорта')
    filename = fields.Char('Имя файла')

    # Параметры экспорта
    period_id = fields.Many2one('budget.period', 'Период', required=True)
    company_ids = fields.Many2many('res.company', string='Предприятия')
    include_all_companies = fields.Boolean('Включить все предприятия', default=True)

    # Настройки формата
    sheet_name = fields.Char('Название листа', default='Бюджет 24')
    include_history = fields.Boolean('Включить исторические данные', default=True)
    currency_display = fields.Selection([
        ('thousands', 'тыс. грн'),
        ('units', 'грн'),
        ('thousands_usd', 'тыс. USD')
    ], 'Отображение валюты', default='thousands')

    # Результаты
    result_file = fields.Binary('Результирующий файл', readonly=True)
    result_filename = fields.Char('Имя результирующего файла', readonly=True)
    import_summary = fields.Text('Результат импорта', readonly=True)

    def action_import_investment_budget(self):
        """Импорт инвестиционного бюджета из Excel"""
        if not self.import_file:
            raise UserError('Загрузите Excel файл!')

        try:
            # Парсим Excel файл
            projects_data = self._parse_investment_excel()

            # Создаем или находим инвестиционный бюджет
            budget_type = self.env['budget.type'].search([('code', '=', '24')], limit=1)
            if not budget_type:
                raise UserError('Тип бюджета "24 - Инвестиционный" не найден!')

            # Определяем ЦБО группы компаний
            holding_cbo = self.env['budget.responsibility.center'].search([
                ('cbo_type', '=', 'holding')
            ], limit=1)

            if not holding_cbo:
                raise UserError('ЦБО холдинга не найдено!')

            # Создаем основной бюджет
            budget = self.env['budget.plan'].create({
                'period_id': self.period_id.id,
                'budget_type_id': budget_type.id,
                'cbo_id': holding_cbo.id,
                'responsible_user_id': self.env.user.id,
                'state': 'draft',
                'notes': f'Импортировано из файла: {self.filename}'
            })

            # Создаем инвестиционные проекты и связанные бюджетные линии
            imported_projects = 0
            total_amount = 0

            for project_data in projects_data:
                project = self._create_investment_project(project_data)
                budget_line = self._create_budget_line_from_project(budget, project, project_data)
                imported_projects += 1
                total_amount += project_data.get('total_amount', 0)

            self.import_summary = f"""
Импорт инвестиционного бюджета завершен!

Статистика:
- Импортировано проектов: {imported_projects}
- Общая сумма: {total_amount:,.2f} тыс.грн
- Бюджет ID: {budget.id}
- Статус: {budget.state}

Структура проектов:
- Основные проекты: {len([p for p in projects_data if p.get('level') == 1])}
- Подпроекты: {len([p for p in projects_data if p.get('level') > 1])}

✅ Готово к планированию и утверждению!
            """

            return {
                'type': 'ir.actions.act_window',
                'name': 'Импортированный инвестиционный бюджет',
                'res_model': 'budget.plan',
                'res_id': budget.id,
                'view_mode': 'form',
                'target': 'current'
            }

        except Exception as e:
            raise UserError(f'Ошибка импорта: {str(e)}')

    def action_export_investment_budget(self):
        """Экспорт инвестиционных бюджетов в Excel"""
        if not OPENPYXL_AVAILABLE:
            raise UserError('Библиотека openpyxl не установлена!')

        try:
            # Собираем данные бюджетов
            budget_data = self._collect_investment_budget_data()

            # Создаем Excel файл
            excel_file = self._create_investment_excel(budget_data)

            # Сохраняем результат
            filename = f"Инвест_ГК_{self.period_id.name}_{datetime.now().strftime('%Y%m%d')}.xlsx"
            self.result_file = base64.b64encode(excel_file)
            self.result_filename = filename

            return {
                'type': 'ir.actions.act_window',
                'res_model': 'investment.budget.wizard',
                'res_id': self.id,
                'view_mode': 'form',
                'target': 'new',
                'context': {'show_result': True}
            }

        except Exception as e:
            raise UserError(f'Ошибка экспорта: {str(e)}')

    def _parse_investment_excel(self):
        """Парсинг Excel файла инвестиционного бюджета"""
        file_content = base64.b64decode(self.import_file)
        file_stream = io.BytesIO(file_content)

        workbook = load_workbook(file_stream, data_only=True)
        worksheet = workbook['Бюджет 24']  # Основной лист

        projects_data = []
        current_row = 8  # Начинаем с 8 строки (после заголовков)

        while current_row <= worksheet.max_row:
            project_data = self._parse_project_row(worksheet, current_row)
            if project_data:
                projects_data.append(project_data)
            current_row += 1

            # Защита от бесконечного цикла
            if current_row > 1000:
                break

        return projects_data

    def _parse_project_row(self, worksheet, row_num):
        """Парсинг одной строки проекта"""
        try:
            # Основные поля
            project_num = worksheet[f'A{row_num}'].value
            project_name = worksheet[f'B{row_num}'].value
            total_amount = worksheet[f'C{row_num}'].value
            paid_2023 = worksheet[f'D{row_num}'].value or 0
            paid_2024 = worksheet[f'E{row_num}'].value or 0
            remaining = worksheet[f'F{row_num}'].value or 0

            if not project_name:
                return None

            # Определяем уровень проекта по номеру
            level = 1
            if isinstance(project_num, str):
                if '.' in project_num:
                    level = project_num.count('.') + 1

            # Месячная разбивка (колонки G-R)
            monthly_amounts = {}
            months = ['january', 'february', 'march', 'april', 'may', 'june',
                      'july', 'august', 'september', 'october', 'november', 'december']

            for i, month in enumerate(months):
                col_letter = chr(ord('G') + i)  # G, H, I, ...
                amount = worksheet[f'{col_letter}{row_num}'].value or 0
                monthly_amounts[month] = float(amount) if amount else 0

            return {
                'project_number': project_num,
                'name': str(project_name).strip(),
                'total_amount': float(total_amount) if total_amount else 0,
                'paid_2023': float(paid_2023),
                'paid_2024': float(paid_2024),
                'remaining_amount': float(remaining),
                'monthly_amounts': monthly_amounts,
                'level': level,
                'row_number': row_num
            }

        except Exception as e:
            _logger.warning(f"Ошибка парсинга строки {row_num}: {e}")
            return None

    def _create_investment_project(self, project_data):
        """Создание инвестиционного проекта"""
        # Создаем проект в модуле project
        project = self.env['project.project'].create({
            'name': project_data['name'],
            'is_sales_project': False,  # Это инвестиционный проект
            'partner_id': self.env.company.partner_id.id,
            'company_id': self.env.company.id,
            'date_start': self.period_id.date_start,
            'description': f"""
Инвестиционный проект из бюджета
Номер: {project_data.get('project_number', 'N/A')}
Общая сумма: {project_data['total_amount']:,.2f} тыс.грн
Оплачено в 2023: {project_data['paid_2023']:,.2f} тыс.грн
Оплачено в 2024: {project_data['paid_2024']:,.2f} тыс.грн
Остаток: {project_data['remaining_amount']:,.2f} тыс.грн
            """
        })

        # Создаем задачи для каждого месяца с планируемыми суммами
        for month, amount in project_data['monthly_amounts'].items():
            if amount > 0:
                self.env['project.task'].create({
                    'name': f"{project_data['name']} - {month.title()}",
                    'project_id': project.id,
                    'description': f"Планируемая сумма: {amount:,.2f} тыс.грн",
                    'company_id': self.env.company.id,
                })

        return project

    def _create_budget_line_from_project(self, budget, project, project_data):
        """Создание строки бюджета на основе проекта"""
        return self.env['budget.plan.line'].create({
            'plan_id': budget.id,
            'description': project_data['name'],
            'project_id': project.id,
            'planned_amount': project_data['total_amount'] * 1000,  # Переводим в грн
            'quantity': 1,
            'unit_price': project_data['total_amount'] * 1000,
            'calculation_method': 'manual',
            'calculation_basis': f"Инвестиционный проект {project_data.get('project_number', '')}",
            'notes': f"Остаток к доплате: {project_data['remaining_amount']:,.2f} тыс.грн"
        })

    def _collect_investment_budget_data(self):
        """Сбор данных инвестиционных бюджетов для экспорта"""
        # Находим все инвестиционные бюджеты за период
        budget_type = self.env['budget.type'].search([('code', '=', '24')], limit=1)

        domain = [
            ('period_id', '=', self.period_id.id),
            ('budget_type_id', '=', budget_type.id),
            ('state', 'in', ['approved', 'executed'])
        ]

        if not self.include_all_companies:
            domain.append(('company_ids', 'in', self.company_ids.ids))

        budgets = self.env['budget.plan'].search(domain)

        # Группируем данные по компаниям и проектам
        companies_data = {}

        for budget in budgets:
            main_company = budget.company_ids[0] if budget.company_ids else None
            company_key = main_company.name if main_company.company_ids else None
            if company_key not in companies_data:
                companies_data[company_key] = {
                    'company': main_company,
                    'total_amount': 0,
                    'projects': []
                }

            # Добавляем проекты из бюджетных линий
            for line in budget.line_ids:
                if line.project_id:
                    project_data = self._extract_project_data(line)
                    companies_data[company_key]['projects'].append(project_data)
                    companies_data[company_key]['total_amount'] += project_data['total_amount']

        return companies_data

    def _extract_project_data(self, budget_line):
        """Извлечение данных проекта из бюджетной линии"""
        project = budget_line.project_id

        # Получаем месячную разбивку из задач проекта
        monthly_amounts = {month: 0 for month in [
            'january', 'february', 'march', 'april', 'may', 'june',
            'july', 'august', 'september', 'october', 'november', 'december'
        ]}

        # Простое равномерное распределение (можно улучшить)
        monthly_amount = budget_line.planned_amount / 12000  # Переводим в тыс.грн и делим на 12
        for month in monthly_amounts:
            monthly_amounts[month] = monthly_amount

        return {
            'name': budget_line.description,
            'total_amount': budget_line.planned_amount / 1000,  # Переводим в тыс.грн
            'paid_2023': 0,  # Можно получить из фактических данных
            'paid_2024': budget_line.actual_amount / 1000,
            'remaining_amount': (budget_line.planned_amount - budget_line.actual_amount) / 1000,
            'monthly_amounts': monthly_amounts,
            'project_number': project.id if project else '',
        }

    def _create_investment_excel(self, companies_data):
        """Создание Excel файла в формате ГК Хлебодар"""
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = self.sheet_name

        # Стили
        header_font = Font(bold=True, size=12)
        border = Border(left=Side(style='thin'), right=Side(style='thin'),
                        top=Side(style='thin'), bottom=Side(style='thin'))

        # Заголовок документа
        worksheet['A1'] = f'Месяц планирования:'
        worksheet['B1'] = self.period_id.name
        worksheet['D1'] = 'Сумма всех бюджетов'

        total_sum = sum(data['total_amount'] for data in companies_data.values())
        worksheet['E1'] = total_sum

        worksheet['A2'] = 'тыс. грн. с НДС'

        # Заголовки таблицы
        headers = [
            '№ пп', 'Название проекта', 'Сумма проекта',
            'Оплачено в 2023 году', 'Оплачено в 2024 году', 'Остаток платежей',
            'Январь', 'Февраль', 'Март', 'Апрель', 'Май', 'Июнь',
            'Июль', 'Август', 'Сентябрь', 'Октябрь', 'Ноябрь', 'Декабрь'
        ]

        for col, header in enumerate(headers, 1):
            cell = worksheet.cell(row=3, column=col)
            cell.value = header
            cell.font = header_font
            cell.border = border

        # Заполняем данные по компаниям
        current_row = 4

        for company_name, company_data in companies_data.items():
            # Заголовок компании
            worksheet.cell(row=current_row, column=1, value=company_name).font = header_font
            worksheet.cell(row=current_row, column=3, value=company_data['total_amount'])
            current_row += 1

            # Проекты компании
            for i, project in enumerate(company_data['projects'], 1):
                worksheet.cell(row=current_row, column=1, value=f"{i}.")
                worksheet.cell(row=current_row, column=2, value=project['name'])
                worksheet.cell(row=current_row, column=3, value=project['total_amount'])
                worksheet.cell(row=current_row, column=4, value=project['paid_2023'])
                worksheet.cell(row=current_row, column=5, value=project['paid_2024'])
                worksheet.cell(row=current_row, column=6, value=project['remaining_amount'])

                # Месячная разбивка
                for col, month in enumerate(['january', 'february', 'march', 'april', 'may', 'june',
                                             'july', 'august', 'september', 'october', 'november', 'december'], 7):
                    worksheet.cell(row=current_row, column=col, value=project['monthly_amounts'][month])

                current_row += 1

            current_row += 1  # Пустая строка между компаниями

        # Автоширина колонок
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width

        # Сохраняем в память
        excel_buffer = io.BytesIO()
        workbook.save(excel_buffer)
        excel_buffer.seek(0)

        return excel_buffer.read()