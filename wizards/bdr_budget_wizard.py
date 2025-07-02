# -*- coding: utf-8 -*-

import base64
import io
import logging
from datetime import datetime, date
from odoo import models, fields, api
from odoo.exceptions import UserError, ValidationError

_logger = logging.getLogger('budget.wizard')

try:
    import openpyxl
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, NamedStyle

    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


class BdrBudgetWizard(models.TransientModel):
    """Майстер роботи з БДР (Бюджет доходів і витрат)"""
    _name = 'bdr.budget.wizard'
    _description = 'Імпорт/Експорт БДР'

    operation_type = fields.Selection([
        ('import', 'Імпорт з Excel'),
        ('export', 'Експорт в Excel'),
        ('analyze', 'Аналіз даних')
    ], 'Тип операції', required=True, default='import')

    # Параметри імпорту
    import_file = fields.Binary('Excel файл БДР', help="Завантажте файл БДР у форматі Excel")
    filename = fields.Char('Ім\'я файлу')

    # Параметри бюджету
    period_id = fields.Many2one('budget.period', 'Період', required=True)
    company_id = fields.Many2one('res.company', 'Підприємство',
                                 default=lambda self: self.env.company, required=True)
    cbo_id = fields.Many2one('budget.responsibility.center', 'ЦБО', required=True)

    # Налаштування структури БДР
    sheet_name = fields.Char('Назва аркуша', default='БДР',
                             help="Назва аркуша з даними БДР")
    start_row = fields.Integer('Початковий рядок даних', default=3,
                               help="Рядок з якого починаються дані БДР")

    # Структура колонок БДР
    item_column = fields.Char('Колонка статей', default='A',
                              help="Колонка з назвами статей БДР")
    jan_column = fields.Char('Колонка Січень', default='B')
    feb_column = fields.Char('Колонка Лютий', default='C')
    mar_column = fields.Char('Колонка Березень', default='D')
    apr_column = fields.Char('Колонка Квітень', default='E')
    may_column = fields.Char('Колонка Травень', default='F')
    jun_column = fields.Char('Колонка Червень', default='G')
    jul_column = fields.Char('Колонка Липень', default='H')
    aug_column = fields.Char('Колонка Серпень', default='I')
    sep_column = fields.Char('Колонка Вересень', default='J')
    oct_column = fields.Char('Колонка Жовтень', default='K')
    nov_column = fields.Char('Колонка Листопад', default='L')
    dec_column = fields.Char('Колонка Грудень', default='M')
    total_column = fields.Char('Колонка Разом', default='N')

    # Налаштування обробки
    create_monthly_budgets = fields.Boolean('Створити помісячні бюджети', default=True,
                                            help="Створити окремі бюджети для кожного місяця")
    update_existing = fields.Boolean('Оновити існуючі', default=False,
                                     help="Оновити існуючі бюджети замість створення нових")
    auto_categorize = fields.Boolean('Автоматична категоризація', default=True,
                                     help="Автоматично визначати категорії витрат")

    # Валютні налаштування
    currency_multiplier = fields.Selection([
        ('1', 'грн'),
        ('1000', 'тис. грн'),
        ('1000000', 'млн. грн')
    ], 'Множник валюти', default='1000',
        help="У якому форматі представлені суми в БДР")

    # Результати
    result_file = fields.Binary('Результуючий файл', readonly=True)
    result_filename = fields.Char('Ім\'я результуючого файлу', readonly=True)
    import_summary = fields.Text('Результат імпорту', readonly=True)
    preview_data = fields.Text('Попередній перегляд', readonly=True)

    # Статистика
    total_income = fields.Monetary('Загальні доходи', readonly=True, currency_field='currency_id')
    total_expenses = fields.Monetary('Загальні витрати', readonly=True, currency_field='currency_id')
    net_profit = fields.Monetary('Чистий прибуток', readonly=True, currency_field='currency_id')
    currency_id = fields.Many2one('res.currency', related='company_id.currency_id', readonly=True)
    income_growth = fields.Float('Зростання доходів (%)', compute='_compute_bdr_analysis', store=False)
    profit_margin = fields.Float('Маржа прибутку (%)', compute='_compute_bdr_analysis', store=False)

    @api.depends('period_id', 'company_id', 'cbo_id', 'operation_type')
    def _compute_bdr_analysis(self):
        """Обчислення показників БДР для аналізу"""
        for wizard in self:
            if wizard.operation_type != 'analyze' or not wizard.period_id:
                wizard.total_income = 0
                wizard.total_expenses = 0
                wizard.net_profit = 0
                wizard.income_growth = 0
                wizard.profit_margin = 0
                continue

            # Пошук бюджетів доходів
            income_domain = [
                ('period_id', '=', wizard.period_id.id),
                ('budget_type_id.budget_category', '=', 'income'),
                ('state', 'in', ['approved', 'executed'])
            ]
            if wizard.company_id:
                income_domain.append(('company_id', '=', wizard.company_id.id))
            if wizard.cbo_id:
                income_domain.append(('cbo_id', '=', wizard.cbo_id.id))

            income_budgets = self.env['budget.plan'].search(income_domain)
            wizard.total_income = sum(income_budgets.mapped('planned_amount'))

            # Пошук бюджетів витрат
            expense_domain = [
                ('period_id', '=', wizard.period_id.id),
                ('budget_type_id.budget_category', 'in', ['direct_costs', 'indirect_costs', 'administrative']),
                ('state', 'in', ['approved', 'executed'])
            ]
            if wizard.company_id:
                expense_domain.append(('company_id', '=', wizard.company_id.id))
            if wizard.cbo_id:
                expense_domain.append(('cbo_id', '=', wizard.cbo_id.id))

            expense_budgets = self.env['budget.plan'].search(expense_domain)
            wizard.total_expenses = sum(expense_budgets.mapped('planned_amount'))

            # Розрахунок прибутку та маржі
            wizard.net_profit = wizard.total_income - wizard.total_expenses
            wizard.profit_margin = (wizard.net_profit / wizard.total_income * 100) if wizard.total_income != 0 else 0

            # Розрахунок зростання

            previous_period = self.env['budget.period'].search([
                ('date_start', '<', wizard.period_id.date_start),
                ('company_id', '=', wizard.company_id.id if wizard.company_id else False)
            ], order='date_start desc', limit=1)

            if previous_period:
                # Пошук доходів попереднього періоду
                prev_income_domain = [
                    ('period_id', '=', previous_period.id),
                    ('budget_type_id.budget_category', '=', 'income'),
                    ('state', 'in', ['approved', 'executed'])
                ]
                if wizard.company_id:
                    prev_income_domain.append(('company_id', '=', wizard.company_id.id))
                if wizard.cbo_id:
                    prev_income_domain.append(('cbo_id', '=', wizard.cbo_id.id))

                prev_income_budgets = self.env['budget.plan'].search(prev_income_domain)
                prev_total_income = sum(prev_income_budgets.mapped('planned_amount'))

                # Розрахунок відсотка зростання
                if prev_total_income > 0:
                    wizard.income_growth = ((wizard.total_income - prev_total_income) / prev_total_income) * 100
                else:
                    wizard.income_growth = 100 if wizard.total_income > 0 else 0
            else:
                wizard.income_growth = 0

    def action_preview_bdr(self):
        """Попередній перегляд даних БДР"""
        if not self.import_file:
            raise UserError('Завантажте файл БДР!')

        try:
            bdr_data = self._parse_bdr_excel()

            # Формуємо попередній перегляд
            preview_lines = []
            total_income = 0
            total_expenses = 0

            for section_name, section_data in bdr_data.items():
                preview_lines.append(f"\n=== {section_name.upper()} ===")

                for item in section_data['items'][:5]:  # Показуємо перші 5 позицій
                    annual_total = sum(item['monthly_amounts'].values())
                    preview_lines.append(f"{item['name']}: {annual_total:,.2f}")

                    if 'дохід' in section_name.lower() or 'виручка' in section_name.lower():
                        total_income += annual_total
                    else:
                        total_expenses += annual_total

                if len(section_data['items']) > 5:
                    preview_lines.append(f"... та ще {len(section_data['items']) - 5} позицій")

            # Додаємо підсумки
            preview_lines.append(f"\n=== ПІДСУМКИ ===")
            preview_lines.append(f"Загальні доходи: {total_income:,.2f}")
            preview_lines.append(f"Загальні витрати: {total_expenses:,.2f}")
            preview_lines.append(f"Чистий прибуток: {total_income - total_expenses:,.2f}")

            self.preview_data = '\n'.join(preview_lines)

            # Оновлюємо статистику
            multiplier = float(self.currency_multiplier)
            self.total_income = total_income * multiplier
            self.total_expenses = total_expenses * multiplier
            self.net_profit = (total_income - total_expenses) * multiplier

            return {
                'type': 'ir.actions.act_window',
                'res_model': 'bdr.budget.wizard',
                'res_id': self.id,
                'view_mode': 'form',
                'target': 'new',
                'context': {'preview_mode': True}
            }

        except Exception as e:
            raise UserError(f'Помилка аналізу файлу БДР: {str(e)}')

    def action_import_bdr(self):
        """Імпорт БДР в систему бюджетування"""
        if not self.import_file:
            raise UserError('Завантажте файл БДР!')

        try:
            # Парсимо дані БДР
            bdr_data = self._parse_bdr_excel()

            created_budgets = []
            imported_lines = 0

            # Створюємо бюджети по розділах БДР
            for section_name, section_data in bdr_data.items():
                # Визначаємо тип бюджету
                budget_type = self._determine_budget_type(section_name)

                if not budget_type:
                    _logger.warning(f"Не знайдено тип бюджету для розділу: {section_name}")
                    continue

                # Створюємо основний бюджет розділу
                section_budget = self._create_section_budget(section_name, budget_type, section_data)
                created_budgets.append(section_budget)

                # Створюємо позиції бюджету
                for item in section_data['items']:
                    if self.create_monthly_budgets:
                        imported_lines += self._create_monthly_budget_lines(section_budget, item)
                    else:
                        imported_lines += self._create_annual_budget_line(section_budget, item)

                # Створюємо помісячні бюджети якщо потрібно
                if self.create_monthly_budgets:
                    monthly_budgets = self._create_monthly_budgets(section_budget, section_data)
                    created_budgets.extend(monthly_budgets)

            # Формуємо звіт
            self.import_summary = f"""
✅ Імпорт БДР завершено успішно!

📊 Статистика:
• Розділів БДР: {len(bdr_data)}
• Створено бюджетів: {len(created_budgets)}
• Імпортовано позицій: {imported_lines}
• Загальні доходи: {self.total_income:,.2f} {self.currency_id.name}
• Загальні витрати: {self.total_expenses:,.2f} {self.currency_id.name}
• Чистий прибуток: {self.net_profit:,.2f} {self.currency_id.name}

🏗️ Створені бюджети:
{chr(10).join([f"• {b.display_name}" for b in created_budgets[:10]])}
{f"... та ще {len(created_budgets) - 10} бюджетів" if len(created_budgets) > 10 else ""}

🎯 Готово до планування та затвердження!
            """

            # Повертаємо список створених бюджетів
            return {
                'type': 'ir.actions.act_window',
                'name': 'Імпортовані бюджети БДР',
                'res_model': 'budget.plan',
                'view_mode': 'tree,form',
                'domain': [('id', 'in', [b.id for b in created_budgets])],
                'context': {'default_period_id': self.period_id.id}
            }

        except Exception as e:
            raise UserError(f'Помилка імпорту БДР: {str(e)}')

    def action_export_bdr(self):
        """Експорт бюджетів у формат БДР"""
        if not OPENPYXL_AVAILABLE:
            raise UserError('Бібліотека openpyxl не встановлена!')

        try:
            # Збираємо дані бюджетів
            budget_data = self._collect_budget_data_for_bdr()

            # Створюємо Excel файл БДР
            excel_file = self._create_bdr_excel(budget_data)

            # Зберігаємо результат
            filename = f"БДР_{self.company_id.name}_{self.period_id.name}_{datetime.now().strftime('%Y%m%d')}.xlsx"
            self.result_file = base64.b64encode(excel_file)
            self.result_filename = filename

            return {
                'type': 'ir.actions.act_window',
                'res_model': 'bdr.budget.wizard',
                'res_id': self.id,
                'view_mode': 'form',
                'target': 'new',
                'context': {'show_result': True}
            }

        except Exception as e:
            raise UserError(f'Помилка експорту БДР: {str(e)}')

    def _parse_bdr_item(self, worksheet, row_num, month_columns):
        """Парсинг одної позиції БДР"""
        try:
            item_name = self._get_cell_value(worksheet, row_num, self.item_column)
            if not item_name:
                return None

            # Читаємо помісячні дані
            monthly_amounts = {}
            annual_total = 0

            for month, column in month_columns.items():
                amount = self._get_cell_value(worksheet, row_num, column)
                try:
                    amount = float(amount) if amount else 0.0
                except (ValueError, TypeError):
                    amount = 0.0

                monthly_amounts[month] = amount
                annual_total += amount

            # Перевіряємо загальну суму з колонки "Разом"
            total_from_file = self._get_cell_value(worksheet, row_num, self.total_column)
            if total_from_file:
                try:
                    total_from_file = float(total_from_file)
                    # Якщо є розбіжність, використовуємо суму з файлу
                    if abs(annual_total - total_from_file) > 0.01:
                        _logger.warning(f"Розбіжність у сумах для {item_name}: {annual_total} vs {total_from_file}")
                        annual_total = total_from_file
                except:
                    pass

            return {
                'name': str(item_name).strip(),
                'monthly_amounts': monthly_amounts,
                'annual_total': annual_total,
                'row_number': row_num
            }

        except Exception as e:
            _logger.error(f"Помилка парсингу рядка {row_num}: {e}")
            return None

    def _parse_bdr_excel(self):
        """Парсинг Excel файлу БДР"""
        if not OPENPYXL_AVAILABLE:
            raise UserError('Бібліотека openpyxl не встановлена!')

        file_content = base64.b64decode(self.import_file)
        file_stream = io.BytesIO(file_content)

        try:
            workbook = load_workbook(file_stream, data_only=True)

            # Вибираємо аркуш
            if self.sheet_name in workbook.sheetnames:
                worksheet = workbook[self.sheet_name]
            else:
                worksheet = workbook.active

            bdr_data = {}
            current_section = None
            current_row = self.start_row

            # Колонки місяців
            month_columns = {
                'january': self.jan_column,
                'february': self.feb_column,
                'march': self.mar_column,
                'april': self.apr_column,
                'may': self.may_column,
                'june': self.jun_column,
                'july': self.jul_column,
                'august': self.aug_column,
                'september': self.sep_column,
                'october': self.oct_column,
                'november': self.nov_column,
                'december': self.dec_column
            }

            while current_row <= worksheet.max_row:
                item_name = self._get_cell_value(worksheet, current_row, self.item_column)

                if not item_name:
                    current_row += 1
                    continue

                item_name = str(item_name).strip()

                # Визначаємо чи це заголовок розділу
                if self._is_section_header(item_name):
                    current_section = item_name
                    if current_section not in bdr_data:
                        bdr_data[current_section] = {'items': [], 'total': 0}
                    current_row += 1
                    continue

                # Парсимо дані позиції
                if current_section:
                    item_data = self._parse_bdr_item(worksheet, current_row, month_columns)
                    if item_data:
                        bdr_data[current_section]['items'].append(item_data)
                        bdr_data[current_section]['total'] += sum(item_data['monthly_amounts'].values())

                current_row += 1

                # Захист від нескінченного циклу
                if current_row > self.start_row + 1000:
                    break

            return bdr_data

        finally:
            file_stream.close()



    def _is_section_header(self, text):
        """Визначення чи є текст заголовком розділу БДР"""
        section_keywords = [
            'дохід', 'виручка', 'надходження',
            'собівартість', 'витрати', 'затрати',
            'валовий', 'операційні', 'адміністративні',
            'фінансові', 'прибуток', 'збиток'
        ]

        text_lower = text.lower()
        return any(keyword in text_lower for keyword in section_keywords)

    def _determine_budget_type(self, section_name):
        """Визначення типу бюджету за назвою розділу БДР"""
        section_lower = section_name.lower()

        # Маппінг розділів БДР до типів бюджетів
        section_mapping = {
            'дохід': '09',  # Бюджет доходів по реалізації послуг
            'виручка': '09',
            'собівартість': '11',  # Бюджет витрат на логістику
            'витрати': '01',  # Бюджет ФОТ (за замовчуванням)
            'адміністративні': '07',  # Бюджет консультаційних послуг
            'фінансові': '08',  # Бюджет фінансової діяльності
            'маркетинг': '05',  # Бюджет маркетингових витрат
        }

        for keyword, budget_code in section_mapping.items():
            if keyword in section_lower:
                budget_type = self.env['budget.type'].search([('code', '=', budget_code)], limit=1)
                if budget_type:
                    return budget_type

        # За замовчуванням повертаємо загальний тип
        return self.env['budget.type'].search([('code', '=', '01')], limit=1)

    def _create_section_budget(self, section_name, budget_type, section_data):
        """Створення бюджету для розділу БДР"""
        return self.env['budget.plan'].create({
            'period_id': self.period_id.id,
            'budget_type_id': budget_type.id,
            'cbo_id': self.cbo_id.id,
            'company_id': self.company_id.id,
            'responsible_user_id': self.env.user.id,
            'state': 'draft',
            'notes': f'Імпортовано з БДР: {section_name}\nФайл: {self.filename}'
        })

    def _create_annual_budget_line(self, budget, item_data):
        """Створення річної позиції бюджету"""
        # Автоматична категоризація
        category = None
        if self.auto_categorize:
            category = self._auto_determine_category(item_data['name'])

        multiplier = float(self.currency_multiplier)

        line_vals = {
            'plan_id': budget.id,
            'description': item_data['name'],
            'planned_amount': item_data['annual_total'] * multiplier,
            'quantity': 12,  # 12 місяців
            'unit_price': (item_data['annual_total'] * multiplier) / 12,
            'calculation_method': 'manual',
            'calculation_basis': f'Імпорт з БДР, рядок {item_data["row_number"]}',
        }

        if category:
            line_vals['budget_category_id'] = category.id

        self.env['budget.plan.line'].create(line_vals)
        return 1

    def _create_monthly_budget_lines(self, budget, item_data):
        """Створення помісячних позицій бюджету"""
        multiplier = float(self.currency_multiplier)
        lines_created = 0

        month_names = {
            'january': 'Січень', 'february': 'Лютий', 'march': 'Березень',
            'april': 'Квітень', 'may': 'Травень', 'june': 'Червень',
            'july': 'Липень', 'august': 'Серпень', 'september': 'Вересень',
            'october': 'Жовтень', 'november': 'Листопад', 'december': 'Грудень'
        }

        # Автоматична категоризація
        category = None
        if self.auto_categorize:
            category = self._auto_determine_category(item_data['name'])

        for month_eng, amount in item_data['monthly_amounts'].items():
            if amount != 0:  # Створюємо тільки ненульові позиції
                month_ukr = month_names.get(month_eng, month_eng)

                line_vals = {
                    'plan_id': budget.id,
                    'description': f"{item_data['name']} ({month_ukr})",
                    'planned_amount': amount * multiplier,
                    'quantity': 1,
                    'unit_price': amount * multiplier,
                    'calculation_method': 'manual',
                    'calculation_basis': f'БДР {month_ukr}, рядок {item_data["row_number"]}',
                }

                if category:
                    line_vals['budget_category_id'] = category.id

                self.env['budget.plan.line'].create(line_vals)
                lines_created += 1

        return lines_created

    def _auto_determine_category(self, item_name):
        """Автоматичне визначення категорії витрат"""
        item_lower = item_name.lower()

        category_keywords = {
            'зарплата': 'SALARY',
            'премія': 'SALARY',
            'фот': 'SALARY',
            'комунальні': 'UTILITIES',
            'електроенергія': 'UTILITIES',
            'газ': 'UTILITIES',
            'вода': 'UTILITIES',
            'зв\'язок': 'COMM',
            'телефон': 'COMM',
            'інтернет': 'COMM',
            'програма': 'SOFTWARE',
            'ліцензія': 'SOFTWARE',
            'ремонт': 'MAINT',
            'обслуговування': 'MAINT',
            'запчастини': 'SPARE',
            'канцтовари': 'OFFICE',
            'відрядження': 'TRAVEL',
            'навчання': 'TRAINING',
            'реклама': 'MARKETING',
            'маркетинг': 'MARKETING',
        }

        for keyword, category_code in category_keywords.items():
            if keyword in item_lower:
                category = self.env['budget.category'].search([('code', '=', category_code)], limit=1)
                if category:
                    return category

        return None

    def _get_cell_value(self, worksheet, row, column):
        """Отримання значення комірки"""
        try:
            if isinstance(column, str):
                cell = worksheet[f"{column}{row}"]
            else:
                cell = worksheet.cell(row=row, column=column)
            return cell.value
        except:
            return None

    def _collect_budget_data_for_bdr(self):
        """Збір даних бюджетів для експорту в БДР"""
        # Знаходимо всі бюджети за період
        budgets = self.env['budget.plan'].search([
            ('period_id', '=', self.period_id.id),
            ('company_id', '=', self.company_id.id),
            ('state', 'in', ['approved', 'executed'])
        ])

        bdr_data = {}

        for budget in budgets:
            section_name = budget.budget_type_id.name

            if section_name not in bdr_data:
                bdr_data[section_name] = []

            for line in budget.line_ids:
                # Розбиваємо річну суму по місяцях (спрощено)
                monthly_amount = line.planned_amount / 12

                bdr_data[section_name].append({
                    'name': line.description,
                    'monthly_amounts': {
                        month: monthly_amount for month in [
                            'january', 'february', 'march', 'april', 'may', 'june',
                            'july', 'august', 'september', 'october', 'november', 'december'
                        ]
                    },
                    'annual_total': line.planned_amount
                })

        return bdr_data

    def _create_bdr_excel(self, bdr_data):
        """Створення Excel файлу в форматі БДР"""
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "БДР"

        # Стилі
        header_style = NamedStyle(name="header")
        header_style.font = Font(bold=True, size=12)
        header_style.alignment = Alignment(horizontal="center")

        section_style = NamedStyle(name="section")
        section_style.font = Font(bold=True, size=11)

        # Заголовок
        worksheet['A1'] = f'БЮДЖЕТ ДОХОДІВ І ВИТРАТ'
        worksheet['A2'] = f'{self.company_id.name} - {self.period_id.name}'

        # Заголовки колонок
        headers = ['Стаття', 'Січень', 'Лютий', 'Березень', 'Квітень', 'Травень', 'Червень',
                   'Липень', 'Серпень', 'Вересень', 'Жовтень', 'Листопад', 'Грудень', 'Разом']

        for col, header in enumerate(headers, 1):
            cell = worksheet.cell(row=4, column=col)
            cell.value = header
            cell.style = header_style

        current_row = 5

        # Заповнюємо дані
        for section_name, items in bdr_data.items():
            # Заголовок розділу
            cell = worksheet.cell(row=current_row, column=1)
            cell.value = section_name
            cell.style = section_style
            current_row += 1

            # Позиції розділу
            for item in items:
                worksheet.cell(row=current_row, column=1, value=item['name'])

                col = 2
                annual_total = 0
                for month in ['january', 'february', 'march', 'april', 'may', 'june',
                              'july', 'august', 'september', 'october', 'november', 'december']:
                    amount = item['monthly_amounts'].get(month, 0)
                    worksheet.cell(row=current_row, column=col, value=amount)
                    annual_total += amount
                    col += 1

                # Разом
                worksheet.cell(row=current_row, column=14, value=annual_total)
                current_row += 1

            current_row += 1  # Пуста строка між розділами

        # Автоширина колонок
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width

        # Зберігаємо у пам'ять
        excel_buffer = io.BytesIO()
        workbook.save(excel_buffer)
        excel_buffer.seek(0)

        return excel_buffer.read()

    def action_analyze_bdr(self):
        """Аналіз структури БДР"""
        if not self.import_file:
            raise UserError('Завантажте файл БДР для аналізу!')

        try:
            bdr_data = self._parse_bdr_excel()

            analysis_result = []
            analysis_result.append("=== АНАЛІЗ СТРУКТУРИ БДР ===\n")

            total_sections = len(bdr_data)
            total_items = sum(len(section['items']) for section in bdr_data.values())

            analysis_result.append(f"📊 Загальна статистика:")
            analysis_result.append(f"• Розділів: {total_sections}")
            analysis_result.append(f"• Загальна кількість статей: {total_items}")
            analysis_result.append("")

            # Аналіз по розділах
            for section_name, section_data in bdr_data.items():
                analysis_result.append(f"📈 {section_name}:")
                analysis_result.append(f"  • Кількість статей: {len(section_data['items'])}")
                analysis_result.append(f"  • Загальна сума: {section_data['total']:,.2f}")

                if section_data['items']:
                    max_item = max(section_data['items'], key=lambda x: x['annual_total'])
                    analysis_result.append(
                        f"  • Найбільша стаття: {max_item['name']} ({max_item['annual_total']:,.2f})")

                analysis_result.append("")

            # Рекомендації
            analysis_result.append("💡 РЕКОМЕНДАЦІЇ:")

            # Перевірка на помилки
            errors = []

            for section_name, section_data in bdr_data.items():
                for item in section_data['items']:
                    # Перевірка на негативні доходи або позитивні витрати
                    if 'дохід' in section_name.lower() and item['annual_total'] < 0:
                        errors.append(f"❌ Негативний дохід: {item['name']}")

                    # Перевірка на нульові значення
                    if item['annual_total'] == 0:
                        errors.append(f"⚠️ Нульове значення: {item['name']}")

            if errors:
                analysis_result.append("🔍 Виявлені проблеми:")
                analysis_result.extend(errors[:10])  # Показуємо перші 10 помилок
                if len(errors) > 10:
                    analysis_result.append(f"... та ще {len(errors) - 10} проблем")
            else:
                analysis_result.append("✅ Структура БДР коректна")

            analysis_result.append("")
            analysis_result.append("🎯 Готовність до імпорту:")
            analysis_result.append(f"• Можна створити {total_sections} типів бюджетів")
            analysis_result.append(f"• Буде імпортовано {total_items} бюджетних позицій")

            self.import_summary = '\n'.join(analysis_result)

            return {
                'type': 'ir.actions.act_window',
                'res_model': 'bdr.budget.wizard',
                'res_id': self.id,
                'view_mode': 'form',
                'target': 'new',
                'context': {'analysis_mode': True}
            }

        except Exception as e:
            raise UserError(f'Помилка аналізу БДР: {str(e)}')

    def _create_monthly_budgets(self, parent_budget, section_data):
        """Створення окремих помісячних бюджетів"""
        monthly_budgets = []

        month_names = {
            'january': 'Січень', 'february': 'Лютий', 'march': 'Березень',
            'april': 'Квітень', 'may': 'Травень', 'june': 'Червень',
            'july': 'Липень', 'august': 'Серпень', 'september': 'Вересень',
            'october': 'Жовтень', 'november': 'Листопад', 'december': 'Грудень'
        }

        for month_eng, month_ukr in month_names.items():
            # Рахуємо загальну суму місяця
            month_total = sum(item['monthly_amounts'].get(month_eng, 0) for item in section_data['items'])

            if month_total != 0:  # Створюємо тільки ненульові місяці
                # Знаходимо період місяця
                month_period = self._find_or_create_month_period(month_eng, month_ukr)

                if month_period:
                    monthly_budget = self.env['budget.plan'].create({
                        'period_id': month_period.id,
                        'budget_type_id': parent_budget.budget_type_id.id,
                        'cbo_id': parent_budget.cbo_id.id,
                        'company_id': parent_budget.company_id.id,
                        'responsible_user_id': parent_budget.responsible_user_id.id,
                        'parent_budget_id': parent_budget.id,
                        'state': 'draft',
                        'notes': f'Помісячний бюджет з БДР: {month_ukr}\nБатьківський бюджет: {parent_budget.display_name}'
                    })

                    monthly_budgets.append(monthly_budget)

                    # Створюємо позиції місячного бюджету
                    self._create_month_budget_lines(monthly_budget, section_data['items'], month_eng)

        return monthly_budgets

    def _find_or_create_month_period(self, month_eng, month_ukr):
        """Знаходження або створення місячного періоду"""
        # Визначаємо номер місяця
        month_numbers = {
            'january': 1, 'february': 2, 'march': 3, 'april': 4,
            'may': 5, 'june': 6, 'july': 7, 'august': 8,
            'september': 9, 'october': 10, 'november': 11, 'december': 12
        }

        month_num = month_numbers.get(month_eng)
        if not month_num:
            return None

        # Отримуємо рік з основного періоду
        if self.period_id.period_type == 'year':
            year = self.period_id.date_start.year
        else:
            year = self.period_id.date_start.year

        # Шукаємо існуючий місячний період
        start_date = date(year, month_num, 1)

        month_period = self.env['budget.period'].search([
            ('company_id', '=', self.company_id.id),
            ('period_type', '=', 'month'),
            ('date_start', '=', start_date)
        ], limit=1)

        if not month_period:
            # Створюємо новий місячний період
            from dateutil.relativedelta import relativedelta
            end_date = start_date + relativedelta(months=1) - relativedelta(days=1)

            month_period = self.env['budget.period'].create({
                'name': f"{month_ukr} {year}",
                'period_type': 'month',
                'date_start': start_date,
                'date_end': end_date,
                'company_id': self.company_id.id,
                'state': 'draft'
            })

        return month_period

    def _create_month_budget_lines(self, monthly_budget, items, month_eng):
        """Створення позицій місячного бюджету"""
        multiplier = float(self.currency_multiplier)

        for item in items:
            month_amount = item['monthly_amounts'].get(month_eng, 0)

            if month_amount != 0:
                # Автоматична категоризація
                category = None
                if self.auto_categorize:
                    category = self._auto_determine_category(item['name'])

                line_vals = {
                    'plan_id': monthly_budget.id,
                    'description': item['name'],
                    'planned_amount': month_amount * multiplier,
                    'quantity': 1,
                    'unit_price': month_amount * multiplier,
                    'calculation_method': 'manual',
                    'calculation_basis': f'БДР {month_eng}, рядок {item["row_number"]}',
                }

                if category:
                    line_vals['budget_category_id'] = category.id

                self.env['budget.plan.line'].create(line_vals)

    def action_create_bdr_template(self):
        """Створення шаблону БДР для компанії"""
        try:
            # Основні розділи БДР
            bdr_sections = [
                {
                    'name': 'Доходи від реалізації',
                    'budget_type_code': '09',
                    'items': [
                        'Виручка від реалізації продукції',
                        'Доходи від послуг',
                        'Інші операційні доходи'
                    ]
                },
                {
                    'name': 'Прямі витрати',
                    'budget_type_code': '11',
                    'items': [
                        'Сировина та матеріали',
                        'Заробітна плата виробничого персоналу',
                        'Амортизація виробничого обладнання'
                    ]
                },
                {
                    'name': 'Операційні витрати',
                    'budget_type_code': '01',
                    'items': [
                        'Заробітна плата адміністративного персоналу',
                        'Витрати на утримання офісу',
                        'Комунальні послуги'
                    ]
                },
                {
                    'name': 'Маркетингові витрати',
                    'budget_type_code': '05',
                    'items': [
                        'Витрати на рекламу',
                        'Участь у виставках',
                        'Маркетингові дослідження'
                    ]
                }
            ]

            created_templates = []

            for section in bdr_sections:
                # Знаходимо тип бюджету
                budget_type = self.env['budget.type'].search([('code', '=', section['budget_type_code'])], limit=1)
                if not budget_type:
                    continue

                # Створюємо шаблон
                template = self.env['budget.template'].create({
                    'name': f"Шаблон БДР - {section['name']}",
                    'budget_type_id': budget_type.id,
                    'cbo_type': 'enterprise',  # Для підприємств
                    'description': f"Автоматично створений шаблон для розділу БДР: {section['name']}",
                    'is_default': True,
                    'company_id': self.company_id.id
                })

                # Створюємо позиції шаблону
                for i, item_name in enumerate(section['items'], 1):
                    self.env['budget.template.line'].create({
                        'template_id': template.id,
                        'sequence': i * 10,
                        'description': item_name,
                        'default_quantity': 1.0,
                        'default_unit_price': 0.0,
                        'calculation_method': 'manual',
                        'is_mandatory': True,
                        'allow_edit': True,
                        'notes': f'Позиція БДР розділу "{section["name"]}"'
                    })

                created_templates.append(template)

            # Повідомлення про результат
            self.import_summary = f"""
✅ Створено шаблони БДР для {self.company_id.name}

📋 Створені шаблони:
{chr(10).join([f"• {t.name}" for t in created_templates])}

🎯 Використання:
1. Відкрийте розділ "Шаблони бюджетів"
2. Оберіть потрібний шаблон БДР
3. Створіть бюджет на його основі
4. Заповніть планові суми

💡 Шаблони можна редагувати під потреби вашого підприємства
            """

            return {
                'type': 'ir.actions.act_window',
                'name': 'Створені шаблони БДР',
                'res_model': 'budget.template',
                'view_mode': 'tree,form',
                'domain': [('id', 'in', [t.id for t in created_templates])],
                'context': {'default_company_id': self.company_id.id}
            }

        except Exception as e:
            raise UserError(f'Помилка створення шаблонів БДР: {str(e)}')