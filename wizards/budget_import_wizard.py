# -*- coding: utf-8 -*-

import base64
import io
import logging
from odoo import models, fields, api
from odoo.exceptions import UserError, ValidationError

_logger = logging.getLogger('budget.import')

try:
    import openpyxl
    from openpyxl import load_workbook

    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    _logger.warning("openpyxl not available. Excel import will be limited.")

try:
    import pandas as pd

    PANDAS_AVAILABLE = True
except ImportError:
    PANDAS_AVAILABLE = False


class BudgetImportWizard(models.TransientModel):
    """Универсальный мастер импорта бюджетов из Excel"""
    _name = 'budget.import.wizard'
    _description = 'Импорт бюджетов из Excel'

    # Основные параметры
    name = fields.Char('Название импорта', default='Импорт бюджета')
    import_file = fields.Binary('Excel файл', required=True)
    filename = fields.Char('Имя файла')

    # Параметры бюджета
    period_id = fields.Many2one('budget.period', 'Период', required=True)
    budget_type_id = fields.Many2one('budget.type', 'Тип бюджета', required=True)
    cbo_id = fields.Many2one('budget.responsibility.center', 'ЦБО', required=True)
    company_id = fields.Many2one('res.company', 'Компания', default=lambda self: self.env.company)

    # Настройки импорта
    sheet_name = fields.Char('Название листа', help="Оставьте пустым для первого листа")
    start_row = fields.Integer('Начальная строка данных', default=10,
                               help="Строка, с которой начинаются данные (не заголовки)")

    # Маппинг колонок
    description_column = fields.Char('Колонка описания', default='B',
                                     help="Колонка с описанием статьи бюджета")
    amount_column = fields.Char('Колонка суммы', default='E',
                                help="Колонка с плановой суммой")
    quantity_column = fields.Char('Колонка количества', default='C',
                                  help="Колонка с количеством (необязательно)")
    price_column = fields.Char('Колонка цены', default='D',
                               help="Колонка с ценой за единицу (необязательно)")

    # Дополнительные настройки
    skip_empty_rows = fields.Boolean('Пропускать пустые строки', default=True)
    auto_create_accounts = fields.Boolean('Автоматически создавать счета', default=False)
    default_calculation_method = fields.Selection([
        ('manual', 'Ручной ввод'),
        ('norm_based', 'На основе нормативов'),
        ('statistical', 'Статистический'),
        ('contract_based', 'На основе договоров')
    ], 'Метод расчета по умолчанию', default='manual')

    # Результаты предпросмотра
    preview_data = fields.Text('Предпросмотр данных', readonly=True)
    import_summary = fields.Text('Результат импорта', readonly=True)

    # Настройки валидации
    validate_amounts = fields.Boolean('Проверять суммы', default=True)
    min_amount = fields.Float('Минимальная сумма', default=0.01)
    max_amount = fields.Float('Максимальная сумма', default=10000000)

    @api.onchange('import_file', 'filename')
    def _onchange_import_file(self):
        """Автоматическое определение параметров из файла"""
        if self.import_file and self.filename:
            try:
                # Попытка автоматического определения типа бюджета по имени файла
                filename_lower = self.filename.lower()

                # Маппинг типов бюджетов по ключевым словам в названии файла
                budget_type_mapping = {
                    'ит': '12',  # Бюджет информационного обеспечения
                    'фот': '01',  # Бюджет ФОТ
                    'маркетинг': '05',  # Бюджет маркетинговых расходов
                    'энерго': '19',  # Бюджет энергоносителей
                    'инвест': '24',  # Инвестиционный бюджет
                }

                for keyword, code in budget_type_mapping.items():
                    if keyword in filename_lower:
                        budget_type = self.env['budget.type'].search([('code', '=', code)], limit=1)
                        if budget_type:
                            self.budget_type_id = budget_type.id
                            break

                # Попытка определить ЦБО по файлу
                if 'ук' in filename_lower:
                    cbo = self.env['budget.responsibility.center'].search([
                        ('budget_level', '=', 'tactical')
                    ], limit=1)
                    if cbo:
                        self.cbo_id = cbo.id

            except Exception as e:
                _logger.warning(f"Ошибка автоматического определения параметров: {e}")

    def action_preview_data(self):
        """Предпросмотр данных перед импортом"""
        if not self.import_file:
            raise UserError('Загрузите Excel файл!')

        try:
            data = self._parse_excel_file()
            preview_lines = []

            for i, row_data in enumerate(data[:10]):  # Показываем первые 10 строк
                preview_lines.append(f"Строка {i + 1}: {row_data}")

            self.preview_data = '\n'.join(preview_lines)

            if len(data) > 10:
                self.preview_data += f"\n... и еще {len(data) - 10} строк"

            return {
                'type': 'ir.actions.act_window',
                'res_model': 'budget.import.wizard',
                'res_id': self.id,
                'view_mode': 'form',
                'target': 'new',
                'context': {'preview_mode': True}
            }

        except Exception as e:
            raise UserError(f'Ошибка анализа файла: {str(e)}')

    def action_import_budget(self):
        """Выполнение импорта бюджета"""
        if not self.import_file:
            raise UserError('Загрузите Excel файл!')

        try:
            # Проверяем, не существует ли уже такой бюджет
            existing_budget = self.env['budget.plan'].search([
                ('period_id', '=', self.period_id.id),
                ('budget_type_id', '=', self.budget_type_id.id),
                ('cbo_id', '=', self.cbo_id.id)
            ])

            if existing_budget:
                raise UserError(f'Бюджет уже существует: {existing_budget.display_name}')

            # Парсим данные
            data = self._parse_excel_file()

            if not data:
                raise UserError('Нет данных для импорта!')

            # Создаем бюджет
            budget = self._create_budget_plan()

            # Создаем строки бюджета
            imported_lines = 0
            skipped_lines = 0

            for row_data in data:
                try:
                    if self._create_budget_line(budget, row_data):
                        imported_lines += 1
                    else:
                        skipped_lines += 1
                except Exception as e:
                    _logger.warning(f"Ошибка создания строки бюджета: {e}")
                    skipped_lines += 1

            # Обновляем итоги
            budget._compute_totals()

            self.import_summary = f"""
Импорт завершен успешно!

Бюджет: {budget.display_name}
Импортировано строк: {imported_lines}
Пропущено строк: {skipped_lines}
Общая сумма: {budget.planned_amount:,.2f} {budget.currency_id.name}

ID бюджета: {budget.id}
            """

            return {
                'type': 'ir.actions.act_window',
                'name': 'Импортированный бюджет',
                'res_model': 'budget.plan',
                'res_id': budget.id,
                'view_mode': 'form',
                'target': 'current'
            }

        except Exception as e:
            raise UserError(f'Ошибка импорта: {str(e)}')

    def _parse_excel_file(self):
        """Парсинг Excel файла"""
        if not OPENPYXL_AVAILABLE:
            raise UserError('Библиотека openpyxl не установлена. Обратитесь к администратору.')

        # Декодируем файл
        file_content = base64.b64decode(self.import_file)
        file_stream = io.BytesIO(file_content)

        try:
            workbook = load_workbook(file_stream, data_only=True)

            # Выбираем лист
            if self.sheet_name:
                if self.sheet_name in workbook.sheetnames:
                    worksheet = workbook[self.sheet_name]
                else:
                    raise UserError(f'Лист "{self.sheet_name}" не найден в файле!')
            else:
                worksheet = workbook.active

            data = []
            current_row = self.start_row

            while current_row <= worksheet.max_row:
                row_data = self._parse_row(worksheet, current_row)

                if row_data:
                    data.append(row_data)
                elif not self.skip_empty_rows:
                    break

                current_row += 1

                # Защита от бесконечного цикла
                if current_row > self.start_row + 1000:
                    break

            return data

        except Exception as e:
            raise UserError(f'Ошибка чтения Excel файла: {str(e)}')
        finally:
            file_stream.close()

    def _parse_row(self, worksheet, row_num):
        """Парсинг одной строки"""
        try:
            # Получаем значения из указанных колонок
            description = self._get_cell_value(worksheet, row_num, self.description_column)
            amount = self._get_cell_value(worksheet, row_num, self.amount_column)
            quantity = self._get_cell_value(worksheet, row_num, self.quantity_column)
            price = self._get_cell_value(worksheet, row_num, self.price_column)

            # Пропускаем пустые строки
            if not description and not amount:
                return None

            # Конвертируем числовые значения
            try:
                amount = float(amount) if amount else 0.0
                quantity = float(quantity) if quantity else 1.0
                price = float(price) if price else amount  # Если цена не указана, используем общую сумму
            except (ValueError, TypeError):
                amount = 0.0
                quantity = 1.0
                price = 0.0

            # Валидация
            if self.validate_amounts and (amount < self.min_amount or amount > self.max_amount):
                _logger.warning(f"Сумма {amount} в строке {row_num} вне допустимого диапазона")
                return None

            return {
                'description': str(description).strip() if description else f"Позиция {row_num}",
                'planned_amount': amount,
                'quantity': quantity,
                'unit_price': price,
                'row_number': row_num
            }

        except Exception as e:
            _logger.error(f"Ошибка парсинга строки {row_num}: {e}")
            return None

    def _get_cell_value(self, worksheet, row, column):
        """Получение значения ячейки"""
        try:
            if isinstance(column, str):
                cell = worksheet[f"{column}{row}"]
            else:
                cell = worksheet.cell(row=row, column=column)
            return cell.value
        except:
            return None

    def _create_budget_plan(self):
        """Создание основного документа бюджета"""
        return self.env['budget.plan'].create({
            'period_id': self.period_id.id,
            'budget_type_id': self.budget_type_id.id,
            'cbo_id': self.cbo_id.id,
            'company_id': self.company_id.id,
            'responsible_user_id': self.env.user.id,
            'state': 'draft',
            'notes': f'Импортировано из файла: {self.filename}'
        })

    def _create_budget_line(self, budget, row_data):
        """Создание строки бюджета"""
        try:
            # Создаем строку бюджета
            line_vals = {
                'plan_id': budget.id,
                'description': row_data['description'],
                'planned_amount': row_data['planned_amount'],
                'quantity': row_data['quantity'],
                'unit_price': row_data['unit_price'],
                'calculation_method': self.default_calculation_method,
                'calculation_basis': f"Импорт из Excel, строка {row_data['row_number']}"
            }

            # Попытка автоматического определения счета
            account = self._find_account_by_description(row_data['description'])
            if account:
                line_vals['account_id'] = account.id

            self.env['budget.plan.line'].create(line_vals)
            return True

        except Exception as e:
            _logger.error(f"Ошибка создания строки бюджета: {e}")
            return False

    def _find_account_by_description(self, description):
        """Поиск счета по описанию"""
        # Простые правила маппинга для ИТ бюджета
        account_mapping = {
            'запчаст': '2081',  # Запчасти
            'ремонт': '2082',  # Ремонт и обслуживание
            'програм': '2131',  # Программное обеспечение
            'связь': '2140',  # Связь
            'интернет': '2140',  # Интернет
            'мобильн': '2140',  # Мобильная связь
            'командир': '2137',  # Командировки
        }

        description_lower = description.lower()

        for keyword, account_code in account_mapping.items():
            if keyword in description_lower:
                account = self.env['account.account'].search([
                    ('code', 'like', account_code),
                    ('company_id', '=', self.company_id.id)
                ], limit=1)

                if account:
                    return account
                elif self.auto_create_accounts:
                    # Создаем счет автоматически (упрощенная логика)
                    return self._create_account(account_code, description)

        return None

    def _create_account(self, code, description):
        """Создание нового счета (если включена опция)"""
        try:
            return self.env['account.account'].create({
                'code': code,
                'name': f"Автосозданный счет: {description[:50]}",
                'account_type': 'expense',
                'company_id': self.company_id.id
            })
        except:
            return None

    def action_download_template(self):
        """Скачивание шаблона Excel для импорта"""
        # Здесь можно реализовать создание шаблона Excel
        raise UserError('Функция скачивания шаблона будет реализована в следующей версии')

    def action_batch_import(self):
        """Массовый импорт нескольких файлов"""
        # Для будущей реализации пакетного импорта
        raise UserError('Массовый импорт будет реализован в следующей версии')