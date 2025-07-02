# -*- coding: utf-8 -*-
# wizards/budget_category_manager_wizard.py

from odoo import models, fields, api
from odoo.exceptions import UserError, ValidationError
import logging

_logger = logging.getLogger('budget.wizard')


class BudgetCategoryManagerWizard(models.TransientModel):
    """Майстер для управління категоріями БДР - створення, редагування, імпорт"""
    _name = 'budget.category.manager.wizard'
    _description = 'Управління категоріями БДР'
    _inherit = ['mail.thread', 'mail.activity.mixin']

    operation_type = fields.Selection([
        ('create', 'Створити нову категорію'),
        ('edit', 'Редагувати категорію'),
        ('import_xml', 'Імпорт з XML'),
        ('import_excel', 'Імпорт з Excel'),
        ('bulk_create', 'Масове створення')
    ], 'Тип операції', required=True, default='create')

    # Поля для створення/редагування категорії
    category_id = fields.Many2one('budget.category', 'Категорія для редагування')
    code = fields.Char('Код категорії', size=20, help="Наприклад: 1.100, 2.300.1")
    name = fields.Char('Назва категорії', help="Повна назва категорії")
    description = fields.Text('Опис')
    parent_id = fields.Many2one('budget.category', 'Батьківська категорія')
    sequence = fields.Integer('Послідовність', default=10)
    active = fields.Boolean('Активна', default=True)
    update_existing = fields.Boolean('Оновити існуючі', default=False,
                                     help="Оновити існуючі категорії замість пропуску")
    company_id = fields.Many2one('res.company', 'Підприємство',
                                 default=lambda self: self.env.company)
    # Зопоставлення з рахунками
    default_account_id = fields.Many2one('account.account', 'Рахунок за замовчуванням')
    budget_type_ids = fields.Many2many('budget.type', string='Типи бюджетів')

    # Поля для імпорту
    import_file = fields.Binary('Файл для імпорту')
    filename = fields.Char('Ім\'я файлу')

    # Поля для масового створення
    bulk_categories = fields.Text('Категорії для створення',
                                  help="Формат: код|назва|батьківський_код (кожна категорія з нового рядка)")

    # Результати
    result_message = fields.Text('Результат операції', readonly=True)
    created_categories = fields.Many2many('budget.category',
                                          'wizard_created_categories_rel', string='Створені категорії', readonly=True)

    @api.onchange('operation_type')
    def _onchange_operation_type(self):
        """Очищення полів при зміні типу операції"""
        if self.operation_type == 'edit':
            # Для редагування потрібно вибрати категорію
            pass
        else:
            self.category_id = False
            self.code = False
            self.name = False
            self.description = False
            self.parent_id = False

    @api.onchange('category_id')
    def _onchange_category_id(self):
        """Заповнення полів при виборі категорії для редагування"""
        if self.category_id and self.operation_type == 'edit':
            self.code = self.category_id.code
            self.name = self.category_id.name
            self.description = self.category_id.description
            self.parent_id = self.category_id.parent_id
            self.sequence = self.category_id.sequence
            self.active = self.category_id.active
            self.default_account_id = self.category_id.default_account_id
            self.budget_type_ids = self.category_id.budget_type_ids

    def action_execute(self):
        """Виконання обраної операції"""
        try:
            if self.operation_type == 'create':
                return self._create_category()
            elif self.operation_type == 'edit':
                return self._edit_category()
            elif self.operation_type == 'import_xml':
                return self._import_from_xml()
            elif self.operation_type == 'import_excel':
                return self._import_from_excel()
            elif self.operation_type == 'bulk_create':
                return self._bulk_create_categories()
            else:
                raise UserError('Невідомий тип операції!')
        except Exception as e:
            raise UserError(f'Помилка виконання операції: {str(e)}')

    def _create_category(self):
        """Створення нової категорії"""
        if not self.code or not self.name:
            raise UserError('Код та назва категорії є обов\'язковими!')

        # Перевірка унікальності коду
        existing = self.env['budget.category'].search([('code', '=', self.code)])
        if existing:
            raise UserError(f'Категорія з кодом {self.code} вже існує!')

        # Створення категорії
        category_vals = {
            'code': self.code,
            'name': self.name,
            'description': self.description,
            'parent_id': self.parent_id.id if self.parent_id else False,
            'sequence': self.sequence,
            'active': self.active,
            'default_account_id': self.default_account_id.id if self.default_account_id else False,
            'budget_type_ids': [(6, 0, self.budget_type_ids.ids)]
        }

        new_category = self.env['budget.category'].create(category_vals)

        self.result_message = f'Категорію "{new_category.name}" ({new_category.code}) успішно створено!'
        self.created_categories = [(6, 0, [new_category.id])]

        return self._return_result_view()

    def _edit_category(self):
        """Редагування існуючої категорії"""
        if not self.category_id:
            raise UserError('Виберіть категорію для редагування!')

        if not self.code or not self.name:
            raise UserError('Код та назва категорії є обов\'язковими!')

        # Перевірка унікальності коду (крім поточної категорії)
        existing = self.env['budget.category'].search([
            ('code', '=', self.code),
            ('id', '!=', self.category_id.id)
        ])
        if existing:
            raise UserError(f'Категорія з кодом {self.code} вже існує!')

        # Оновлення категорії
        update_vals = {
            'code': self.code,
            'name': self.name,
            'description': self.description,
            'parent_id': self.parent_id.id if self.parent_id else False,
            'sequence': self.sequence,
            'active': self.active,
            'default_account_id': self.default_account_id.id if self.default_account_id else False,
            'budget_type_ids': [(6, 0, self.budget_type_ids.ids)]
        }

        self.category_id.write(update_vals)

        self.result_message = f'Категорію "{self.category_id.name}" ({self.category_id.code}) успішно оновлено!'
        self.created_categories = [(6, 0, [self.category_id.id])]

        return self._return_result_view()

    def _bulk_create_categories(self):
        """Масове створення категорій"""
        if not self.bulk_categories:
            raise UserError('Введіть дані для створення категорій!')

        lines = self.bulk_categories.strip().split('\n')
        created_categories = []
        errors = []

        for line_num, line in enumerate(lines, 1):
            line = line.strip()
            if not line:
                continue

            try:
                parts = line.split('|')
                if len(parts) < 2:
                    errors.append(f'Рядок {line_num}: Неправильний формат (потрібно: код|назва)')
                    continue

                code = parts[0].strip()
                name = parts[1].strip()
                parent_code = parts[2].strip() if len(parts) > 2 else None

                if not code or not name:
                    errors.append(f'Рядок {line_num}: Порожній код або назва')
                    continue

                # Пошук батьківської категорії
                parent_id = False
                if parent_code:
                    parent = self.env['budget.category'].search([('code', '=', parent_code)], limit=1)
                    if parent:
                        parent_id = parent.id
                    else:
                        errors.append(f'Рядок {line_num}: Батьківська категорія {parent_code} не знайдена')
                        continue

                # Перевірка унікальності
                existing = self.env['budget.category'].search([('code', '=', code)])
                if existing:
                    errors.append(f'Рядок {line_num}: Категорія {code} вже існує')
                    continue

                # Створення категорії
                category_vals = {
                    'code': code,
                    'name': name,
                    'parent_id': parent_id,
                    'sequence': int(code.split('.')[0]) * 100 if '.' in code else 10,
                    'active': True,
                }

                new_category = self.env['budget.category'].create(category_vals)
                created_categories.append(new_category.id)

            except Exception as e:
                errors.append(f'Рядок {line_num}: Помилка створення - {str(e)}')

        # Формування результату
        success_count = len(created_categories)
        error_count = len(errors)

        result_parts = [f'Успішно створено: {success_count} категорій']
        if error_count > 0:
            result_parts.append(f'Помилок: {error_count}')
            result_parts.extend(errors[:10])  # Показуємо перші 10 помилок
            if error_count > 10:
                result_parts.append(f'... та ще {error_count - 10} помилок')

        self.result_message = '\n'.join(result_parts)
        self.created_categories = [(6, 0, created_categories)]

        return self._return_result_view()

    def _import_from_excel(self):
        """Імпорт категорій з Excel файлу"""
        if not self.import_file:
            raise UserError('Завантажте Excel файл!')

        try:
            import base64
            import io
            import openpyxl

            # Декодування файлу
            file_data = base64.b64decode(self.import_file)
            file_obj = io.BytesIO(file_data)

            # Читання Excel
            workbook = openpyxl.load_workbook(file_obj)

            # Вибір аркуша (перший або за назвою)
            if 'БДР' in workbook.sheetnames:
                worksheet = workbook['БДР']
            elif 'бюджет БДиР' in workbook.sheetnames:
                worksheet = workbook['бюджет БДиР']
            else:
                worksheet = workbook.active

            created_categories = []
            errors = []

            # Читання даних з Excel
            for row_num, row in enumerate(worksheet.iter_rows(min_row=1), 1):
                try:
                    # Пропуск порожніх рядків
                    if not any(cell.value for cell in row[:4]):
                        continue

                    # Отримання даних
                    code = row[1].value if len(row) > 1 and row[1].value else None
                    name = row[2].value if len(row) > 2 and row[2].value else None

                    # Перевірка формату коду
                    if not code or not isinstance(code, str) or not code.strip():
                        continue

                    code = code.strip()
                    if not code.replace('.', '').replace('-', '').isdigit():
                        continue

                    if not name:
                        continue

                    name = str(name).strip()

                    # Пошук батьківської категорії
                    parent_id = False
                    code_parts = code.split('.')
                    if len(code_parts) > 2 and code_parts[2]:
                        parent_code = f"{code_parts[0]}.{code_parts[1]}."
                        parent = self.env['budget.category'].search([('code', '=', parent_code)], limit=1)
                        if parent:
                            parent_id = parent.id

                    # Перевірка існування
                    existing = self.env['budget.category'].search([('code', '=', code)])
                    if existing:
                        continue

                    # Створення категорії
                    category_vals = {
                        'code': code,
                        'name': name,
                        'parent_id': parent_id,
                        'sequence': int(code_parts[0]) * 100 if code_parts else 10,
                        'active': True,
                    }

                    new_category = self.env['budget.category'].create(category_vals)
                    created_categories.append(new_category.id)

                except Exception as e:
                    errors.append(f'Рядок {row_num}: {str(e)}')

            success_count = len(created_categories)
            error_count = len(errors)

            result_parts = [f'Імпорт завершено!']
            result_parts.append(f'Створено категорій: {success_count}')
            if error_count > 0:
                result_parts.append(f'Помилок: {error_count}')

            self.result_message = '\n'.join(result_parts)
            self.created_categories = [(6, 0, created_categories)]

            return self._return_result_view()

        except ImportError:
            raise UserError('Бібліотека openpyxl не встановлена!')
        except Exception as e:
            raise UserError(f'Помилка імпорту з Excel: {str(e)}')

    def _return_result_view(self):
        """Повернення представлення з результатом"""
        return {
            'type': 'ir.actions.act_window',
            'name': 'Результат операції',
            'res_model': 'budget.category.manager.wizard',
            'res_id': self.id,
            'view_mode': 'form',
            'target': 'new',
            'context': {'show_result': True}
        }

    def action_open_categories(self):
        """Відкриття створених категорій"""
        if not self.created_categories:
            raise UserError('Немає створених категорій для перегляду!')

        return {
            'type': 'ir.actions.act_window',
            'name': 'Створені категорії',
            'res_model': 'budget.category',
            'view_mode': 'tree,form',
            'domain': [('id', 'in', self.created_categories.ids)],
            'target': 'current',
        }

    def action_create_another(self):
        """Створення ще однієї категорії"""
        return {
            'type': 'ir.actions.act_window',
            'name': 'Управління категоріями БДР',
            'res_model': 'budget.category.manager.wizard',
            'view_mode': 'form',
            'target': 'new',
            'context': {'default_operation_type': self.operation_type}
        }

    def _import_from_xml(self):
        """Імпорт категорій з XML файлу"""
        if not self.import_file:
            raise UserError('Завантажте XML файл для імпорту!')

        try:
            import xml.etree.ElementTree as ET
            import base64

            # Декодування файлу
            file_content = base64.b64decode(self.import_file)
            file_text = file_content.decode('utf-8')

            # Парсинг XML
            root = ET.fromstring(file_text)

            imported_categories = []
            categories_data = []

            # Збір даних з XML
            for record in root.findall('.//record[@model="budget.category"]'):
                category_data = {
                    'xml_id': record.get('id', ''),
                    'code': '',
                    'name': '',
                    'parent_xml_id': '',
                    'description': '',
                    'sequence': 10,
                    'active': True
                }

                # Читання полів
                for field in record.findall('field'):
                    field_name = field.get('name')
                    field_value = field.text or ''

                    if field_name == 'code':
                        category_data['code'] = field_value.strip()
                    elif field_name == 'name':
                        category_data['name'] = field_value.strip()
                    elif field_name == 'description':
                        category_data['description'] = field_value.strip()
                    elif field_name == 'sequence':
                        try:
                            category_data['sequence'] = int(field_value)
                        except:
                            category_data['sequence'] = 10
                    elif field_name == 'active':
                        category_data['active'] = field_value.lower() in ['true', '1']
                    elif field_name == 'parent_id':
                        # Якщо це ref, отримуємо XML ID
                        if field.get('ref'):
                            category_data['parent_xml_id'] = field.get('ref')

                if category_data['code'] and category_data['name']:
                    categories_data.append(category_data)

            if not categories_data:
                raise UserError('У XML файлі не знайдено валідних категорій!')

            # Створення категорій (спочатку батьківські, потім дочірні)
            parent_mapping = {}

            # Перший прохід - створюємо категорії без батьків
            for cat_data in categories_data:
                if not cat_data['parent_xml_id']:
                    category = self._create_category_from_data(cat_data)
                    imported_categories.append(category)
                    parent_mapping[cat_data['xml_id']] = category

            # Другий прохід - створюємо дочірні категорії
            for cat_data in categories_data:
                if cat_data['parent_xml_id']:
                    # Знаходимо батьківську категорію
                    parent_category = None

                    # Спочатку шукаємо у створених
                    if cat_data['parent_xml_id'] in parent_mapping:
                        parent_category = parent_mapping[cat_data['parent_xml_id']]
                    else:
                        # Шукаємо в базі по external ID
                        try:
                            parent_category = self.env.ref(cat_data['parent_xml_id'])
                        except:
                            # Шукаємо по коду
                            parent_code = cat_data['parent_xml_id'].split('_')[-1]
                            parent_category = self.env['budget.category'].search([
                                ('code', '=', parent_code)
                            ], limit=1)

                    cat_data['parent_id'] = parent_category.id if parent_category else False
                    category = self._create_category_from_data(cat_data)
                    imported_categories.append(category)
                    parent_mapping[cat_data['xml_id']] = category

            # Формування результату
            self.import_summary = f"""
    ✅ XML імпорт завершено успішно!

    📊 Статистика:
    - Оброблено записів: {len(categories_data)}
    - Створено категорій: {len(imported_categories)}
    - Файл: {self.filename}

    📋 Створені категорії:
    {chr(10).join([f"• [{cat.code}] {cat.name}" for cat in imported_categories[:10]])}
    {f"... та ще {len(imported_categories) - 10} категорій" if len(imported_categories) > 10 else ""}

    🎯 Всі категорії готові до використання!
            """

            # Повертаємо список створених категорій
            return {
                'type': 'ir.actions.act_window',
                'name': 'Імпортовані категорії',
                'res_model': 'budget.category',
                'view_mode': 'tree,form',
                'domain': [('id', 'in', [cat.id for cat in imported_categories])],
                'context': {'default_active': True}
            }

        except ET.ParseError as e:
            raise UserError(f'Помилка парсингу XML: {str(e)}')
        except Exception as e:
            raise UserError(f'Помилка імпорту XML: {str(e)}')

    def _create_category_from_data(self, cat_data):
        """Створення категорії з даних"""
        # Перевіряємо чи існує категорія з таким кодом
        existing = self.env['budget.category'].search([
            ('code', '=', cat_data['code'])
        ], limit=1)

        if existing and not self.update_existing:
            _logger.warning(f"Категорія з кодом {cat_data['code']} вже існує")
            return existing

        vals = {
            'code': cat_data['code'],
            'name': cat_data['name'],
            'description': cat_data.get('description', ''),
            'sequence': cat_data.get('sequence', 10),
            'active': cat_data.get('active', True),
            'company_id': self.company_id.id if self.company_id else False,
        }

        if cat_data.get('parent_id'):
            vals['parent_id'] = cat_data['parent_id']

        if existing and self.update_existing:
            existing.write(vals)
            return existing
        else:
            return self.env['budget.category'].create(vals)