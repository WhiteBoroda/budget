# -*- coding: utf-8 -*-
# wizards/bdr_budget_wizard_enhanced.py

import base64
import io
import logging
from datetime import datetime
from odoo import models, fields
from odoo.exceptions import UserError

_logger = logging.getLogger('budget.wizard')

try:
    import openpyxl
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, NamedStyle

    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


class BdrBudgetWizardEnhanced(models.TransientModel):
    """Покращений майстер роботи з БДР з підтримкою категорій"""
    _name = 'bdr.budget.wizard.enhanced'
    _description = 'Покращений імпорт/експорт БДР з категоріями'

    operation_type = fields.Selection([
        ('import', 'Імпорт з Excel'),
        ('export', 'Експорт в Excel'),
        ('import_categories', 'Імпорт категорій з Excel'),
        ('analyze', 'Аналіз даних')
    ], 'Тип операції', required=True, default='import')

    # Параметри імпорту
    import_file = fields.Binary('Excel файл БДР', help="Завантажте файл БДР у форматі Excel")
    filename = fields.Char('Ім\'я файлу')

    # Параметри бюджету
    period_id = fields.Many2one('budget.period', 'Період', required=True)
    company_id = fields.Many2one('res.company', 'Підприємство',
                                 default=lambda self: self.env.company, required=True)
    cbo_id = fields.Many2one('budget.responsibility.center', 'ЦБО', required=True)

    # Налаштування структури БДР
    sheet_name = fields.Char('Назва аркуша', default='бюджет БДиР',
                             help="Назва аркуша з даними БДР")
    start_row = fields.Integer('Початковий рядок даних', default=6,
                               help="Рядок з якого починаються дані БДР")

    # Режими роботи з категоріями
    use_categories = fields.Boolean('Використовувати категорії', default=True,
                                    help="Автоматично створювати та призначати категорії")
    auto_create_categories = fields.Boolean('Автоматично створювати категорії', default=True)
    import_categories_only = fields.Boolean('Тільки імпорт категорій', default=False)

    # Структура колонок БДР
    code_column = fields.Char('Колонка кодів', default='B',
                              help="Колонка з кодами статей БДР")
    name_column = fields.Char('Колонка назв', default='C',
                              help="Колонка з назвами статей БДР")
    amount_column = fields.Char('Колонка сум', default='G',
                                help="Колонка з сумами")

    # Результати
    result_file = fields.Binary('Результуючий файл', readonly=True)
    result_filename = fields.Char('Ім\'я результуючого файлу', readonly=True)
    import_summary = fields.Text('Результат імпорту', readonly=True)
    created_categories = fields.Many2many('budget.category',
                                          'wizard_bdr_categories_rel', string='Створені категорії', readonly=True)
    created_budget_lines = fields.Many2many('budget.plan.line',
                                            'wizard_bdr_lines_rel', string='Створені позиції бюджету', readonly=True)

    def action_import_bdr(self):
        """Імпорт БДР з Excel з підтримкою категорій"""
        if not self.import_file:
            raise UserError('Завантажте Excel файл!')

        if not OPENPYXL_AVAILABLE:
            raise UserError('Бібліотека openpyxl не встановлена!')

        try:
            # Створення або пошук плану бюджету
            budget_plan = self._get_or_create_budget_plan()

            # Парсинг Excel файлу
            categories, budget_data = self._parse_bdr_excel_enhanced()

            created_categories = []
            created_lines = []

            # Створення категорій (якщо потрібно)
            if self.use_categories and self.auto_create_categories:
                created_categories = self._create_categories_from_data(categories)

            # Створення позицій бюджету (якщо не тільки категорії)
            if not self.import_categories_only:
                created_lines = self._create_budget_lines_with_categories(
                    budget_plan, budget_data, created_categories)

            # Формування результату
            self._prepare_import_result(created_categories, created_lines)

            return self._return_result_view()

        except Exception as e:
            _logger.error(f'Помилка імпорту БДР: {str(e)}', exc_info=True)
            raise UserError(f'Помилка імпорту БДР: {str(e)}')

    def _parse_bdr_excel_enhanced(self):
        """Покращений парсинг Excel з виділенням категорій та даних"""
        file_data = base64.b64decode(self.import_file)
        file_obj = io.BytesIO(file_data)
        workbook = load_workbook(file_obj, data_only=True)

        # Вибір аркуша
        if self.sheet_name in workbook.sheetnames:
            worksheet = workbook[self.sheet_name]
        else:
            worksheet = workbook.active

        categories = []
        budget_data = []

        # Читання даних з Excel
        for row_num, row in enumerate(worksheet.iter_rows(min_row=self.start_row), self.start_row):
            try:
                # Отримання даних з колонок
                code_cell = self._get_cell_value(row, self.code_column)
                name_cell = self._get_cell_value(row, self.name_column)
                amount_cell = self._get_cell_value(row, self.amount_column)

                # Пропуск порожніх рядків
                if not code_cell and not name_cell:
                    continue

                # Перевірка формату коду категорії
                if code_cell and isinstance(code_cell, str) and self._is_category_code(code_cell):
                    # Це категорія
                    name = str(name_cell).strip() if name_cell else code_cell

                    category_data = {
                        'code': code_cell.strip(),
                        'name': name,
                        'row': row_num,
                        'parent_code': self._determine_parent_code(code_cell.strip())
                    }
                    categories.append(category_data)

                    # Якщо є сума - це також бюджетна позиція
                    if amount_cell and self._is_valid_amount(amount_cell):
                        budget_line_data = {
                            'category_code': code_cell.strip(),
                            'description': name,
                            'amount': self._parse_amount(amount_cell),
                            'row': row_num
                        }
                        budget_data.append(budget_line_data)

                elif name_cell and amount_cell and self._is_valid_amount(amount_cell):
                    # Це позиція бюджету без категорії
                    budget_line_data = {
                        'category_code': code_cell.strip() if code_cell else None,
                        'description': str(name_cell).strip(),
                        'amount': self._parse_amount(amount_cell),
                        'row': row_num
                    }
                    budget_data.append(budget_line_data)

            except Exception as e:
                _logger.warning(f'Помилка обробки рядка {row_num}: {str(e)}')
                continue

        _logger.info(f'Знайдено {len(categories)} категорій та {len(budget_data)} позицій бюджету')
        return categories, budget_data

    def _get_cell_value(self, row, column_letter):
        """Отримання значення комірки за літерою колонки"""
        try:
            column_index = ord(column_letter.upper()) - ord('A')
            if column_index < len(row):
                cell = row[column_index]
                return cell.value if cell else None
            return None
        except:
            return None

    def _is_category_code(self, code):
        """Перевірка чи є код кодом категорії"""
        if not code or not isinstance(code, str):
            return False

        code = code.strip()
        # Код категорії має формат X.XXX. або X.XXX.X
        import re
        pattern = r'^\d+\.\d+\.?(\d+)?$'
        return bool(re.match(pattern, code))

    def _determine_parent_code(self, code):
        """Визначення коду батьківської категорії"""
        parts = code.split('.')
        if len(parts) > 2 and parts[2]:
            # Підкатегорія - батьків X.XXX.
            return f"{parts[0]}.{parts[1]}."
        return None

    def _is_valid_amount(self, value):
        """Перевірка чи є значення валідною сумою"""
        if value is None:
            return False

        if isinstance(value, (int, float)):
            return value != 0

        if isinstance(value, str):
            try:
                float(value.replace(',', '.').replace(' ', ''))
                return True
            except:
                return False

        return False

    def _parse_amount(self, value):
        """Парсинг суми з різних форматів"""
        if isinstance(value, (int, float)):
            return float(value)

        if isinstance(value, str):
            # Видалення пробілів та заміна коми на крапку
            cleaned = value.replace(' ', '').replace(',', '.')
            try:
                return float(cleaned)
            except:
                return 0.0

        return 0.0

    def _create_categories_from_data(self, categories_data):
        """Створення категорій з даних Excel"""
        created_categories = []

        # Спочатку створюємо основні категорії (без батьків)
        main_categories = [cat for cat in categories_data if not cat.get('parent_code')]
        for cat_data in main_categories:
            category = self._create_single_category(cat_data)
            if category:
                created_categories.append(category)

        # Потім створюємо підкатегорії
        sub_categories = [cat for cat in categories_data if cat.get('parent_code')]
        for cat_data in sub_categories:
            category = self._create_single_category(cat_data)
            if category:
                created_categories.append(category)

        return created_categories

    def _create_single_category(self, cat_data):
        """Створення однієї категорії"""
        # Перевірка чи існує категорія
        existing = self.env['budget.category'].search([('code', '=', cat_data['code'])])
        if existing:
            return existing

        # Пошук батьківської категорії
        parent_id = False
        if cat_data.get('parent_code'):
            parent = self.env['budget.category'].search([('code', '=', cat_data['parent_code'])], limit=1)
            if parent:
                parent_id = parent.id

        # Створення категорії
        try:
            category_vals = {
                'code': cat_data['code'],
                'name': cat_data['name'],
                'parent_id': parent_id,
                'sequence': int(cat_data['code'].split('.')[0]) * 100,
                'active': True,
                'company_id': self.company_id.id
            }

            category = self.env['budget.category'].create(category_vals)
            _logger.info(f'Створено категорію: {category.code} - {category.name}')
            return category

        except Exception as e:
            _logger.error(f'Помилка створення категорії {cat_data["code"]}: {str(e)}')
            return None

    def _create_budget_lines_with_categories(self, budget_plan, budget_data, created_categories):
        """Створення позицій бюджету з категоріями"""
        created_lines = []

        for line_data in budget_data:
            try:
                # Пошук категорії
                category = None
                if line_data.get('category_code'):
                    category = self.env['budget.category'].search([
                        ('code', '=', line_data['category_code'])
                    ], limit=1)

                # Створення позиції бюджету
                line_vals = {
                    'plan_id': budget_plan.id,
                    'description': line_data['description'],
                    'planned_amount': line_data['amount'],
                    'budget_category_id': category.id if category else False,
                    'calculation_method': 'manual',
                    'calculation_basis': f'Імпортовано з Excel (рядок {line_data["row"]})'
                }

                # Автоматичне призначення рахунку з категорії
                if category and category.default_account_id:
                    line_vals['account_id'] = category.default_account_id.id

                budget_line = self.env['budget.plan.line'].create(line_vals)
                created_lines.append(budget_line)

            except Exception as e:
                _logger.error(f'Помилка створення позиції бюджету: {str(e)}')
                continue

        return created_lines

    def _get_or_create_budget_plan(self):
        """Створення або пошук плану бюджету БДР"""
        # Пошук типу бюджету БДР
        budget_type = self.env['budget.type'].search([
            ('code', 'in', ['BDR', 'БДР'])
        ], limit=1)

        if not budget_type:
            budget_type = self.env['budget.type'].create({
                'name': 'БДР - Бюджет доходів і витрат',
                'code': 'BDR',
                'description': 'Основний бюджет доходів і витрат'
            })

        # Пошук існуючого плану
        existing_plan = self.env['budget.plan'].search([
            ('period_id', '=', self.period_id.id),
            ('budget_type_id', '=', budget_type.id),
            ('cbo_id', '=', self.cbo_id.id),
            ('company_ids', 'in', [self.company_id.id])
        ], limit=1)

        if existing_plan:
            return existing_plan

        # Створення нового плану
        plan_vals = {
            'name': f'БДР {self.period_id.name} - {self.cbo_id.name}',
            'period_id': self.period_id.id,
            'budget_type_id': budget_type.id,
            'cbo_id': self.cbo_id.id,
            'company_ids': [(6, 0, [self.company_id.id])],
            'state': 'draft'
        }

        return self.env['budget.plan'].create(plan_vals)

    def _prepare_import_result(self, created_categories, created_lines):
        """Підготовка результату імпорту"""
        summary_parts = []
        summary_parts.append(f'Імпорт БДР завершено успішно!')
        summary_parts.append(f'Створено категорій: {len(created_categories)}')
        summary_parts.append(f'Створено позицій бюджету: {len(created_lines)}')

        if created_categories:
            summary_parts.append('\nНові категорії:')
            for category in created_categories[:10]:  # Показуємо перші 10
                summary_parts.append(f'  • {category.code} - {category.name}')
            if len(created_categories) > 10:
                summary_parts.append(f'  ... та ще {len(created_categories) - 10} категорій')

        if created_lines:
            total_amount = sum(line.planned_amount for line in created_lines)
            summary_parts.append(f'\nЗагальна сума бюджету: {total_amount:,.2f} грн')

        self.import_summary = '\n'.join(summary_parts)
        self.created_categories = [(6, 0, [c.id for c in created_categories])]
        self.created_budget_lines = [(6, 0, [l.id for l in created_lines])]

    def _return_result_view(self):
        """Повернення представлення з результатом"""
        return {
            'type': 'ir.actions.act_window',
            'name': 'Результат імпорту БДР',
            'res_model': 'bdr.budget.wizard.enhanced',
            'res_id': self.id,
            'view_mode': 'form',
            'target': 'new',
            'context': {'show_result': True}
        }

    def action_export_bdr(self):
        """Експорт БДР в Excel з категоріями"""
        try:
            # Збір даних для експорту
            budget_data = self._collect_budget_data_for_export()

            # Створення Excel файлу
            excel_file = self._create_bdr_excel_with_categories(budget_data)

            # Збереження результату
            filename = f"БДР_{self.company_id.name}_{self.period_id.name}_{datetime.now().strftime('%Y%m%d')}.xlsx"
            self.result_file = base64.b64encode(excel_file)
            self.result_filename = filename

            return self._return_result_view()

        except Exception as e:
            raise UserError(f'Помилка експорту БДР: {str(e)}')

    def _collect_budget_data_for_export(self):
        """Збір даних бюджету для експорту"""
        # Пошук планів бюджету для експорту
        domain = [
            ('period_id', '=', self.period_id.id),
            ('company_id', '=', self.company_id.id),
            ('cbo_id', '=', self.cbo_id.id)
        ]

        budget_plans = self.env['budget.plan'].search(domain)

        data = {
            'categories': [],
            'lines': [],
            'totals': {}
        }

        # Збір категорій
        categories = self.env['budget.category'].search([('active', '=', True)], order='sequence, code')
        for category in categories:
            data['categories'].append({
                'code': category.code,
                'name': category.name,
                'parent_code': category.parent_id.code if category.parent_id else '',
                'level': len(category.code.split('.')) - 1
            })

        # Збір позицій бюджету
        for plan in budget_plans:
            for line in plan.line_ids:
                data['lines'].append({
                    'category_code': line.budget_category_id.code if line.budget_category_id else '',
                    'description': line.description,
                    'planned_amount': line.planned_amount,
                    'actual_amount': line.actual_amount,
                    'plan_name': plan.name
                })

        return data

    def _create_bdr_excel_with_categories(self, data):
        """Створення Excel файлу БДР з категоріями"""
        if not OPENPYXL_AVAILABLE:
            raise UserError('Бібліотека openpyxl не встановлена!')

        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "БДР"

        # Стилі
        header_style = NamedStyle(name="header")
        header_style.font = Font(bold=True, size=12)
        header_style.alignment = Alignment(horizontal='center')
        header_style.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")

        category_style = NamedStyle(name="category")
        category_style.font = Font(bold=True, size=11)
        category_style.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")

        # Заголовок
        worksheet['A1'] = f"БЮДЖЕТ ДОХОДІВ І ВИТРАТ - {self.company_id.name}"
        worksheet['A1'].style = header_style
        worksheet.merge_cells('A1:G1')

        worksheet['A2'] = f"Період: {self.period_id.name}"
        worksheet['A3'] = f"ЦБО: {self.cbo_id.name}"
        worksheet['A4'] = f"Дата створення: {datetime.now().strftime('%d.%m.%Y %H:%M')}"

        # Заголовки колонок
        headers = ['№', 'Код', 'Найменування статті', 'План, тис. грн', 'Факт, тис. грн', 'Відхилення', '%']
        for col, header in enumerate(headers, 1):
            cell = worksheet.cell(row=6, column=col)
            cell.value = header
            cell.style = header_style

        # Заповнення даних
        current_row = 7
        row_number = 1

        # Групування даних по категоріях
        lines_by_category = {}
        for line in data['lines']:
            category_code = line['category_code'] or 'БЕЗ_КАТЕГОРІЇ'
            if category_code not in lines_by_category:
                lines_by_category[category_code] = []
            lines_by_category[category_code].append(line)

        # Виведення категорій та позицій
        for category in data['categories']:
            if category['code'] in lines_by_category:
                # Заголовок категорії
                if category['level'] == 0:  # Основна категорія
                    worksheet.cell(row=current_row, column=2, value=category['code'])
                    worksheet.cell(row=current_row, column=3, value=category['name'])
                    for col in range(2, 8):
                        worksheet.cell(row=current_row, column=col).style = category_style
                    current_row += 1

                # Позиції в категорії
                category_total = 0
                for line in lines_by_category[category['code']]:
                    worksheet.cell(row=current_row, column=1, value=row_number)
                    worksheet.cell(row=current_row, column=2, value=category['code'])
                    worksheet.cell(row=current_row, column=3, value=line['description'])
                    worksheet.cell(row=current_row, column=4, value=line['planned_amount'] / 1000)
                    worksheet.cell(row=current_row, column=5, value=line['actual_amount'] / 1000)

                    # Відхилення
                    deviation = line['actual_amount'] - line['planned_amount']
                    worksheet.cell(row=current_row, column=6, value=deviation / 1000)

                    # Відсоток виконання
                    if line['planned_amount'] != 0:
                        percentage = (line['actual_amount'] / line['planned_amount']) * 100
                        worksheet.cell(row=current_row, column=7, value=f"{percentage:.1f}%")

                    category_total += line['planned_amount']
                    current_row += 1
                    row_number += 1

        # Позиції без категорій
        if 'БЕЗ_КАТЕГОРІЇ' in lines_by_category:
            worksheet.cell(row=current_row, column=3, value="ПОЗИЦІЇ БЕЗ КАТЕГОРІЙ")
            for col in range(2, 8):
                worksheet.cell(row=current_row, column=col).style = category_style
            current_row += 1

            for line in lines_by_category['БЕЗ_КАТЕГОРІЇ']:
                worksheet.cell(row=current_row, column=1, value=row_number)
                worksheet.cell(row=current_row, column=3, value=line['description'])
                worksheet.cell(row=current_row, column=4, value=line['planned_amount'] / 1000)
                worksheet.cell(row=current_row, column=5, value=line['actual_amount'] / 1000)
                current_row += 1
                row_number += 1

        # Автоматичне розширення колонок
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width

        # Збереження в пам'ять
        output = io.BytesIO()
        workbook.save(output)
        output.seek(0)
        return output.getvalue()

    def action_open_created_categories(self):
        """Відкриття створених категорій"""
        if not self.created_categories:
            raise UserError('Немає створених категорій для перегляду!')

        return {
            'type': 'ir.actions.act_window',
            'name': 'Створені категорії БДР',
            'res_model': 'budget.category',
            'view_mode': 'tree,form',
            'domain': [('id', 'in', self.created_categories.ids)],
            'target': 'current',
        }

    def action_open_created_budget_lines(self):
        """Відкриття створених позицій бюджету"""
        if not self.created_budget_lines:
            raise UserError('Немає створених позицій бюджету для перегляду!')

        return {
            'type': 'ir.actions.act_window',
            'name': 'Створені позиції бюджету',
            'res_model': 'budget.plan.line',
            'view_mode': 'tree,form',
            'domain': [('id', 'in', self.created_budget_lines.ids)],
            'target': 'current',
        }

    def action_download_template(self):
        """Завантаження шаблону Excel для БДР"""
        try:
            template_file = self._create_bdr_template()

            filename = f"Шаблон_БДР_{datetime.now().strftime('%Y%m%d')}.xlsx"
            self.result_file = base64.b64encode(template_file)
            self.result_filename = filename

            return self._return_result_view()

        except Exception as e:
            raise UserError(f'Помилка створення шаблону: {str(e)}')

    def _create_bdr_template(self):
        """Створення шаблону Excel для БДР"""
        if not OPENPYXL_AVAILABLE:
            raise UserError('Бібліотека openpyxl не встановлена!')

        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Шаблон БДР"

        # Заголовки та приклади
        worksheet['A1'] = "ШАБЛОН БЮДЖЕТУ ДОХОДІВ І ВИТРАТ"
        worksheet['A3'] = "Інструкція:"
        worksheet['A4'] = "• Колонка B: Код категорії (наприклад: 1.100., 2.300.1)"
        worksheet['A5'] = "• Колонка C: Назва статті бюджету"
        worksheet['A6'] = "• Колонка G: Планова сума (в гривнях)"

        # Заголовки колонок
        headers = ['№', 'Код', 'Найменування статті', 'Примітки', 'Од.вим.', 'Кількість', 'Сума, грн']
        for col, header in enumerate(headers, 1):
            worksheet.cell(row=8, column=col, value=header)

        # Приклади даних
        examples = [
            ['1.000.', 'ДОХОДИ', '', '', '', '', ''],
            ['1.100.', 'Дохід від реалізації продукції', 'Основна діяльність', 'тис.грн', '1', '1000000'],
            ['2.000.', 'ВИТРАТИ НА ВИРОБНИЦТВО', '', '', '', '', ''],
            ['2.100.', 'Сировина та матеріали', 'Основні матеріали', 'тис.грн', '1', '500000'],
            ['2.200.', 'Заробітна плата виробничого персоналу', 'ФОП виробництва', 'тис.грн', '1', '200000']
        ]

        for row_idx, example in enumerate(examples, 9):
            for col_idx, value in enumerate(example, 2):
                worksheet.cell(row=row_idx, column=col_idx, value=value)

        # Автоматичне розширення колонок
        for col in range(1, 8):
            worksheet.column_dimensions[chr(64 + col)].width = 20

        # Збереження в пам'ять
        output = io.BytesIO()
        workbook.save(output)
        output.seek(0)
        return output.getvalue()